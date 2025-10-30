#!/usr/bin/env python3
"""
Export node-list.xlsx sheets to individual NetAssets-compatible CSV files.

Author: Claude Code Expert System
Date: 2025-10-29
"""

import openpyxl
import csv
import os
from datetime import datetime

def export_sheet_to_csv(wb, sheet_name, output_dir):
    """
    Export a single Excel sheet to CSV format.
    """
    ws = wb[sheet_name]

    # Create safe filename
    safe_name = sheet_name.replace(' ', '_').replace('/', '_')
    output_file = os.path.join(output_dir, f'node-list-{safe_name}.csv')

    rows_exported = 0

    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        for row in ws.iter_rows(values_only=True):
            # Convert datetime objects to strings
            cleaned_row = []
            for cell in row:
                if isinstance(cell, datetime):
                    cleaned_row.append(cell.strftime('%m/%d/%Y %H:%M'))
                elif cell is None:
                    cleaned_row.append('')
                else:
                    cleaned_row.append(str(cell))

            writer.writerow(cleaned_row)
            rows_exported += 1

    return output_file, rows_exported


def main():
    input_file = '/home/michael/Projects/Claude/NetAssets/node-list.xlsx'
    output_dir = '/home/michael/Projects/Claude/NetAssets/node-list-exports'

    print("=" * 80)
    print("📦 NODE-LIST.XLSX EXPORT TOOL")
    print("=" * 80)
    print(f"Input:  {input_file}")
    print(f"Output: {output_dir}/")
    print()

    # Create output directory
    os.makedirs(output_dir, exist_ok=True)

    # Load workbook
    print("📂 Loading Excel workbook...")
    wb = openpyxl.load_workbook(input_file, data_only=True)

    print(f"✅ Found {len(wb.sheetnames)} sheets\n")

    # Export each sheet
    total_rows = 0
    exported_files = []

    for i, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        row_count = ws.max_row
        col_count = ws.max_column

        print(f"  [{i:2d}/{len(wb.sheetnames)}] {sheet_name:25s} ({row_count:4d} rows, {col_count:2d} cols)", end='')

        if row_count <= 1:
            print(" → SKIPPED (empty)")
            continue

        output_file, rows_exported = export_sheet_to_csv(wb, sheet_name, output_dir)
        exported_files.append((sheet_name, output_file, rows_exported))
        total_rows += rows_exported

        print(f" → ✓ {rows_exported} rows")

    print("\n" + "=" * 80)
    print("✅ EXPORT COMPLETE")
    print("=" * 80)
    print(f"\n📊 Summary:")
    print(f"  Total sheets processed: {len(exported_files)}")
    print(f"  Total rows exported:    {total_rows:,}")
    print(f"  Output directory:       {output_dir}")
    print()

    # Show largest files
    print("📈 Largest datasets:")
    sorted_files = sorted(exported_files, key=lambda x: x[2], reverse=True)
    for sheet, file, rows in sorted_files[:5]:
        print(f"  • {sheet:25s} - {rows:4d} rows")

    print()
    print(f"💡 Next step: Import CSV files from {output_dir}/ into NetAssets")
    print()


if __name__ == '__main__':
    main()
